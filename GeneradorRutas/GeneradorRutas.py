from openpyxl import load_workbook
import pandas as pd

archivoCargue = pd.read_excel('FEHCA_CIRCUITO_DPTO_ACCESO_IMP_CODIGO.xlsx')
archivoCodigo = pd.read_excel('CODIGOS_CIRCUITO.xlsx')

wb = load_workbook('FEHCA_CIRCUITO_DPTO_ACCESO_IMP_CODIGO.xlsx')
datos = wb['DATOS_RUTA_TRABAJO']
fecha= input("FECHA AAAAMMDD :")
condicion_No= input("ACCESO_IMP - 0   CIERRE_CTO - 1  :")


if condicion_No == 0:
    condicion = "ACCESO_IMP"
else:
    condicion = " CIERRE_CTO"
    
seguir = "S"
while (seguir == "S") or (seguir == "s") :
    circuito = "none"
    codigo = 0
    departamento = "none"
    NoRuta = "none"
    nombreKml = "none"
    
    circuito = input("CIRCUITO : ")
    tipo_tx = input("TIPO DE RUTA --- AE , SP : ")
    NoRuta = input("Numero de Ruta : ")
    # nombreKml = input("Nombre del kml : ")
    for i in archivoCodigo.index:
        if archivoCodigo["CIRCUITO (ID 147)"][i] == circuito:
            codigo = int(archivoCodigo["ID"][i])
            departamento = archivoCodigo["TERRITORIO (ID 145)"][i]
    
    if codigo == 0:
        codigo = input("CODIGO : ")
        departamento = input("TERRITORIO : ")
    
    xcircuito = circuito.replace(' ', '_') #remplazamos
    nombreKml = xcircuito
    
    
    
    datos.cell(row=2,column=1).value = 'MATRICULACION_' + str(codigo) 
    datos.cell(row=2,column=2).value = "MATRICULACIÓN"
    datos.cell(row=2,column=3).value = str(codigo)
    datos.cell(row=2,column=6).value = str(nombreKml)
   
    if (tipo_tx == "AE") or (tipo_tx == "ae") :
        
        NOMBRE_RUTA = str(xcircuito)+"_"+str(departamento)+"_AE_"+condicion+"_R00"+str(NoRuta)
        
        datos.cell(row=2,column=4).value = NOMBRE_RUTA
        datos.cell(row=2,column=5).value = NOMBRE_RUTA
        
        wb.save(fecha+'_'+xcircuito+'_'+departamento+"_AE_"+condicion+'_'+str(codigo)+'.xlsx')
        
        print('MATRICULACION_' + str(codigo))
        print("MATRICULACIÓN")
        print(str(codigo))
        print(NOMBRE_RUTA)
        print(NOMBRE_RUTA)
        print(str(nombreKml))
    
    elif (tipo_tx == "SP") or (tipo_tx == "sp") :
        
        NOMBRE_RUTA = str(xcircuito)+"_"+str(departamento)+"_SP_"+condicion+"_R00"+str(NoRuta)
        
        datos.cell(row=2,column=4).value = NOMBRE_RUTA
        datos.cell(row=2,column=5).value = NOMBRE_RUTA
        
        wb.save(fecha+'_'+xcircuito+'_'+departamento+"_SP_"+condicion+'_'+str(codigo)+'.xlsx')
        
        print('MATRICULACION_' + str(codigo))
        print("MATRICULACIÓN")
        print(str(codigo))
        print(NOMBRE_RUTA)
        print(NOMBRE_RUTA)
        print(str(nombreKml))
    
    else:
        
        datos.cell(row=2,column=4).value = str(xcircuito)+"_"+str(departamento)+"_SP_"+condicion+"_R00"+str(NoRuta)
        datos.cell(row=2,column=5).value = str(xcircuito)+"_"+str(departamento)+"_SP_"+condicion+"_R00"+str(NoRuta)    

        print('MATRICULACION_' + str(codigo))
        print("MATRICULACIÓN")
        print(str(codigo))
        print(NOMBRE_RUTA)
        print(NOMBRE_RUTA)
        print(str(nombreKml))
        
        wb.save(fecha+'_'+xcircuito+'_'+departamento+"_SP_"+condicion+'_'+str(codigo)+'.xlsx')
    
        datos.cell(row=2,column=4).value = str(xcircuito)+"_"+str(departamento)+"_AE_"+condicion+"_R00"+str(NoRuta)
        datos.cell(row=2,column=5).value = str(xcircuito)+"_"+str(departamento)+"_AE_"+condicion+"_R00"+str(NoRuta)    

        print('MATRICULACION_' + str(codigo))
        print("MATRICULACIÓN")
        print(str(codigo))
        print(NOMBRE_RUTA)
        print(NOMBRE_RUTA)
        print(str(nombreKml))
        
        wb.save(fecha+'_'+xcircuito+'_'+departamento+"_AE_"+condicion+'_'+str(codigo)+'.xlsx')
    
    seguir = input("crear otra ruta : ")


    
